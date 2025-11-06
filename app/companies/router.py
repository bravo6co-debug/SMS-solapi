from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from app.database import get_db
from app.schemas import CompanyCreate, CompanyUpdate, CompanyResponse, CompanyBulkUploadResult
from app.companies import service
from app.auth.router import get_current_user
import openpyxl
from io import BytesIO

router = APIRouter(prefix="/api/companies", tags=["companies"])


@router.get("/test")
def test_endpoint():
    """테스트 엔드포인트 (인증 없음)"""
    return {"message": "테스트 성공", "status": "ok"}


@router.get("", response_model=List[CompanyResponse])
def get_companies(
    search: Optional[str] = Query(None, description="검색어 (이름 또는 아이디)"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=200),
    db: Session = Depends(get_db)
    # current_user = Depends(get_current_user)  # 임시로 비활성화
):
    """발주사 목록 조회"""
    companies = service.get_companies(db, search=search, skip=skip, limit=limit)
    return companies


@router.get("/count")
def get_companies_count(
    db: Session = Depends(get_db)
    # current_user = Depends(get_current_user)  # 임시로 비활성화
):
    """전체 발주사 수"""
    count = service.get_companies_count(db)
    return {"count": count}


@router.get("/{company_id}", response_model=CompanyResponse)
def get_company(
    company_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """발주사 상세 조회"""
    company = service.get_company_by_id(db, company_id)
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="발주사를 찾을 수 없습니다"
        )
    return company


@router.post("", response_model=CompanyResponse, status_code=status.HTTP_201_CREATED)
def create_company(
    company_data: CompanyCreate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """발주사 등록"""
    company = service.create_company(db, company_data)
    return company


@router.put("/{company_id}", response_model=CompanyResponse)
def update_company(
    company_id: int,
    company_data: CompanyUpdate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """발주사 수정"""
    company = service.update_company(db, company_id, company_data)
    if not company:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="발주사를 찾을 수 없습니다"
        )
    return company


@router.delete("/{company_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_company(
    company_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """발주사 삭제"""
    success, error_message = service.delete_company(db, company_id)
    if not success:
        if error_message and "발송 이력" in error_message:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=error_message
            )
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=error_message or "발주사를 찾을 수 없습니다"
        )
    return None


@router.get("/template/download")
def download_template(current_user = Depends(get_current_user)):
    """엑셀 템플릿 다운로드"""
    # 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발주사 등록"

    # 헤더 작성
    headers = ["발주사명*", "전화번호*", "발주사아이디*", "메모"]
    ws.append(headers)

    # 헤더 스타일 적용
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill

    # 예시 데이터 추가
    example_data = ["ABC회사", "01012345678", "ABC001", "메모 예시"]
    ws.append(example_data)

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30

    # BytesIO로 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 파일 다운로드 응답
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=company_upload_template.xlsx"}
    )


@router.post("/bulk-upload", response_model=CompanyBulkUploadResult)
async def bulk_upload_companies(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
    # current_user = Depends(get_current_user)  # 임시로 비활성화
):
    """엑셀 파일로 발주사 대량 등록"""

    # 파일 확장자 체크
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="엑셀 파일만 업로드 가능합니다 (.xlsx, .xls)"
        )

    # 파일 크기 체크 (5MB)
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="파일 크기는 5MB를 초과할 수 없습니다"
        )

    try:
        print(f"[DEBUG] 엑셀 파일 처리 시작: {file.filename}")
        
        # 엑셀 파일 읽기
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active

        # 데이터 추출
        companies_data = []
        max_rows = 501  # 헤더 1행 + 데이터 500행

        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=max_rows, values_only=True)):
            # 빈 행 스킵
            if not any(row):
                continue

            # 최대 500개 제한
            if len(companies_data) >= 500:
                break

            name = str(row[0]).strip() if row[0] else ""
            phone = str(row[1]).strip() if row[1] else ""
            company_id = str(row[2]).strip() if row[2] else ""
            memo = str(row[3]).strip() if row[3] and row[3] != "None" else ""

            companies_data.append({
                "name": name,
                "phone": phone,
                "company_id": company_id,
                "memo": memo
            })

        print(f"[DEBUG] 엑셀에서 추출된 데이터: {len(companies_data)}개")

        if not companies_data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="엑셀 파일에 등록할 데이터가 없습니다"
            )

        # 대량 등록 처리
        print(f"[DEBUG] 대량 등록 서비스 호출 시작")
        success_count, errors = service.create_companies_bulk(db, companies_data)
        print(f"[DEBUG] 대량 등록 서비스 완료: 성공={success_count}, 에러={len(errors)}")

        return CompanyBulkUploadResult(
            success_count=success_count,
            error_count=len(errors),
            errors=errors
        )

    except openpyxl.utils.exceptions.InvalidFileException:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="유효하지 않은 엑셀 파일입니다"
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"파일 처리 중 오류가 발생했습니다: {str(e)}"
        )